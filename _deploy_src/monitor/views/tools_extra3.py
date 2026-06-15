"""ForensicLab 확장 도구 세트 3 — 30개 추가 도구"""
import base64
import datetime as _dt
import io
import json
import os
import re
import socket
import struct
import zipfile
import zlib
import hashlib
import math
from collections import Counter, defaultdict
from pathlib import Path

from flask import request, render_template, jsonify
from monitor.views.tools import bp, _save_log


# ============================================================
# 다중 파일 처리 헬퍼 — 모든 도구 공통
# ============================================================
def get_files(field='file'):
    """request.files 에서 다중 파일 리스트 반환"""
    files = request.files.getlist(field)
    return [f for f in files if f and f.filename]


# ============================================================
# /tools/plist — macOS plist 파서
# ============================================================
def _parse_plist(data: bytes) -> dict:
    if data[:8] == b'bplist00':
        # Binary plist
        try:
            import plistlib
            obj = plistlib.loads(data)
            return {'format': 'Binary (bplist00)', 'data': obj}
        except Exception as e:
            raise ValueError(f'bplist 파싱 실패: {e}')
    elif b'<?xml' in data[:200] and b'<plist' in data[:1000]:
        try:
            import plistlib
            obj = plistlib.loads(data)
            return {'format': 'XML plist', 'data': obj}
        except Exception as e:
            raise ValueError(f'XML plist 파싱 실패: {e}')
    raise ValueError('plist 시그니처 없음 (bplist00 또는 <?xml... <plist)')

def _serialize_plist(obj, depth=0, max_depth=10):
    """plist 결과를 JSON-safe 형태로 변환"""
    if depth > max_depth: return f'... (depth>{max_depth})'
    if isinstance(obj, dict):
        return {str(k): _serialize_plist(v, depth+1, max_depth) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_serialize_plist(v, depth+1, max_depth) for v in obj[:200]]
    if isinstance(obj, bytes):
        return f'<bytes {len(obj)}: {obj[:64].hex()}>'
    if isinstance(obj, _dt.datetime):
        return obj.isoformat()
    return obj

@bp.route('/plist', methods=['GET','POST'])
def plist_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '파일 필요'
        else:
            results = []
            for f in files:
                data = f.read()
                try:
                    r = _parse_plist(data)
                    r['filename'] = f.filename
                    r['data_json'] = _serialize_plist(r['data'])
                    results.append(r)
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/plist.html', result=result, error=error)


# ============================================================
# /tools/amcache — AmCache.hve 전용 파서
# ============================================================
def _parse_amcache(data: bytes) -> dict:
    if data[:4] != b'regf': raise ValueError('regf 시그니처 아님')
    try:
        from Registry.Registry import Registry
        import tempfile
        tf = tempfile.NamedTemporaryFile(delete=False, suffix='.hve')
        tf.write(data); tf.close()
        reg = Registry(tf.name)
        apps = []
        # 최신: Root\InventoryApplicationFile
        try:
            key = reg.open('Root\\InventoryApplicationFile')
            for sub in key.subkeys()[:1000]:
                app = {'subkey': sub.name(), 'last_write': str(sub.timestamp())}
                for val in sub.values():
                    app[val.name()] = str(val.value())[:300]
                apps.append(app)
        except Exception: pass
        # Legacy: Root\File\{VolumeGUID}\<MFTRef>
        legacy = []
        try:
            file_key = reg.open('Root\\File')
            for vol in file_key.subkeys()[:50]:
                for entry in vol.subkeys()[:200]:
                    e = {'volume': vol.name(), 'mft_ref': entry.name()}
                    for val in entry.values():
                        e[val.name()] = str(val.value())[:300]
                    legacy.append(e)
        except Exception: pass
        os.unlink(tf.name)
        return {'modern': apps, 'legacy': legacy,
                'modern_count': len(apps), 'legacy_count': len(legacy)}
    except ImportError:
        raise ValueError('python-registry 라이브러리 필요')

@bp.route('/amcache', methods=['GET','POST'])
def amcache_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '파일 필요'
        else:
            results = []
            for f in files:
                try:
                    r = _parse_amcache(f.read())
                    r['filename'] = f.filename
                    results.append(r)
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/amcache.html', result=result, error=error)


# ============================================================
# /tools/har — HAR 파일 분석
# ============================================================
@bp.route('/har', methods=['GET','POST'])
def har_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '.har 파일 필요'
        else:
            all_files = []
            for f in files:
                try:
                    har = json.loads(f.read())
                    entries = har.get('log', {}).get('entries', [])
                    requests_list = []
                    methods = Counter(); statuses = Counter(); hosts = Counter()
                    total_bytes = 0; total_time = 0
                    for e in entries[:1000]:
                        req = e.get('request', {})
                        resp = e.get('response', {})
                        url = req.get('url', '')
                        host = re.match(r'https?://([^/]+)', url)
                        host = host.group(1) if host else '?'
                        method = req.get('method', '?')
                        status = resp.get('status', 0)
                        sz = resp.get('bodySize', 0)
                        t = e.get('time', 0)
                        methods[method] += 1; statuses[status] += 1; hosts[host] += 1
                        total_bytes += max(sz, 0); total_time += max(t, 0)
                        requests_list.append({
                            'time': e.get('startedDateTime',''),
                            'method': method, 'url': url[:200],
                            'status': status, 'size': sz,
                            'duration': round(t, 1),
                            'mime': resp.get('content',{}).get('mimeType',''),
                        })
                    all_files.append({
                        'filename': f.filename, 'total': len(entries),
                        'requests': requests_list[:200],
                        'methods': methods.most_common(),
                        'statuses': statuses.most_common(),
                        'top_hosts': hosts.most_common(20),
                        'total_bytes': total_bytes,
                        'total_time_ms': round(total_time, 0),
                    })
                except Exception as e:
                    all_files.append({'filename': f.filename, 'error': str(e)})
            result = {'files': all_files}
    return render_template('tools/har.html', result=result, error=error)


# ============================================================
# /tools/sigma — Sigma 규칙 매처
# ============================================================
def _sigma_match(events: list, sigma_yaml: str) -> list:
    """간소화 Sigma: detection.selection 의 key:value 매칭"""
    # YAML 파싱 (간단)
    rule_name = (re.search(r'title:\s*(.+)', sigma_yaml) or [None, '제목없음'])[1].strip()
    desc = (re.search(r'description:\s*(.+)', sigma_yaml) or [None, ''])[1].strip()
    level = (re.search(r'level:\s*(.+)', sigma_yaml) or [None, ''])[1].strip()
    # selection: 블록에서 key: value 추출
    detection_block = re.search(r'detection:\s*\n((?:[ \t].+\n)+)', sigma_yaml)
    if not detection_block: return []
    selection = {}
    for m in re.finditer(r'\s{4,}(\w+):\s*[\'"]?([^\'"\n]+)[\'"]?', detection_block.group(1)):
        key, val = m.group(1), m.group(2).strip()
        selection[key] = val
    matches = []
    for ev in events[:5000]:
        ok = True
        for k, v in selection.items():
            ev_val = str(ev.get(k, '') or ev.get(k.lower(), ''))
            if v.lower() not in ev_val.lower():
                ok = False; break
        if ok:
            matches.append({'event': ev, 'rule': rule_name})
    return [{'rule': rule_name, 'description': desc, 'level': level,
             'matches': matches, 'match_count': len(matches)}]

@bp.route('/sigma', methods=['GET','POST'])
def sigma_tool():
    result = error = None
    if request.method == 'POST':
        sigma_yaml = (request.form.get('sigma') or '').strip()
        events_json = (request.form.get('events') or '').strip()
        if not sigma_yaml: error = 'Sigma 규칙 필요'
        elif not events_json: error = 'JSON 이벤트 배열 필요'
        else:
            try:
                events = json.loads(events_json)
                if not isinstance(events, list): events = [events]
                matches = _sigma_match(events, sigma_yaml)
                result = {'sigma': sigma_yaml, 'events_count': len(events),
                          'matches': matches}
            except Exception as e: error = str(e)
    return render_template('tools/sigma.html', result=result, error=error)


# ============================================================
# /tools/psdeobf — PowerShell 디오브푸스케이션
# ============================================================
def _deobf_ps(text: str) -> dict:
    steps = []
    current = text
    # 1. Base64 디코드
    b64_pattern = re.compile(r"FromBase64String\(['\"]([A-Za-z0-9+/=]+)['\"]\)", re.I)
    for m in b64_pattern.finditer(current):
        try:
            decoded = base64.b64decode(m.group(1)).decode('utf-16-le', errors='ignore')
            if not decoded.strip():
                decoded = base64.b64decode(m.group(1)).decode('utf-8', errors='ignore')
            current = current.replace(m.group(0), f'"{decoded}"')
            steps.append({'step': 'Base64 디코드', 'before': m.group(1)[:60], 'after': decoded[:200]})
        except Exception: pass
    # 2. -EncodedCommand 자동 디코드
    enc_pattern = re.compile(r'-[Ee][Cc][^\s]*\s+([A-Za-z0-9+/=]{20,})')
    for m in enc_pattern.finditer(current):
        try:
            decoded = base64.b64decode(m.group(1)).decode('utf-16-le', errors='ignore')
            current = current.replace(m.group(0), f'-Command "{decoded}"')
            steps.append({'step': '-EncodedCommand', 'before': m.group(1)[:60], 'after': decoded[:200]})
        except Exception: pass
    # 3. 문자열 연결 (+ 'a' + 'b')
    cat_pattern = re.compile(r"['\"]([^'\"]{1,50})['\"]\s*\+\s*['\"]([^'\"]{1,50})['\"]")
    iter_n = 0
    while iter_n < 20:
        new = cat_pattern.sub(lambda m: f'"{m.group(1)}{m.group(2)}"', current)
        if new == current: break
        current = new; iter_n += 1
    if iter_n: steps.append({'step': f'문자열 연결 {iter_n}회', 'before': '', 'after': ''})
    # 4. char 배열 ('A','B','C') → "ABC"
    char_pattern = re.compile(r"\[char\[\]\]\s*\(((?:\d+\s*,\s*)+\d+)\)", re.I)
    for m in char_pattern.finditer(current):
        try:
            nums = [int(x.strip()) for x in m.group(1).split(',')]
            chars = ''.join(chr(n) for n in nums if 0 < n < 128)
            current = current.replace(m.group(0), f'"{chars}"')
            steps.append({'step': 'char 배열', 'before': m.group(1)[:60], 'after': chars[:100]})
        except Exception: pass
    # 5. 백틱 제거 (PowerShell 백틱 이스케이프)
    if '`' in current:
        current_nb = current.replace('`', '')
        if current_nb != current:
            current = current_nb
            steps.append({'step': '백틱(`) 제거', 'before': '', 'after': ''})
    # 6. -Join 연산자
    join_pattern = re.compile(r"\(([^)]+)\)\s*-[Jj]oin\s*['\"]?([^'\"\s]*)['\"]?")
    # IoC 추출
    iocs = {
        'urls': list(set(re.findall(r'https?://[^\s\'"<>]{4,200}', current)))[:20],
        'ips': list(set(re.findall(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', current)))[:20],
        'commands': list(set(re.findall(r'(?:Invoke-Expression|IEX|Invoke-Command|Start-Process|powershell|cmd\.exe|wmic|net\s+user|whoami|hostname)', current, re.I)))[:20],
        'files': list(set(re.findall(r'[A-Za-z]:\\\\?[^\s\'"<>]{3,100}', current)))[:20],
    }
    return {'steps': steps, 'result': current[:50000], 'iocs': iocs,
            'reduction': f'{len(text)} → {len(current)} bytes'}

@bp.route('/psdeobf', methods=['GET','POST'])
def psdeobf_tool():
    result = error = None
    text = ''
    if request.method == 'POST':
        text = (request.form.get('text') or '').strip()
        if not text:
            f = request.files.get('file')
            if f and f.filename: text = f.read().decode('utf-8', errors='replace')
        if not text: error = 'PowerShell 코드 입력 또는 파일 업로드 필요'
        else:
            result = _deobf_ps(text)
            result['original'] = text[:5000]
    return render_template('tools/psdeobf.html', result=result, error=error, text=text)


# ============================================================
# /tools/ioc — IOC 추출기
# ============================================================
_IOC_PATTERNS = {
    'ipv4':     r'\b(?:25[0-5]|2[0-4]\d|[01]?\d\d?)(?:\.(?:25[0-5]|2[0-4]\d|[01]?\d\d?)){3}\b',
    'ipv6':     r'\b(?:[0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}\b',
    'cidr':     r'\b(?:\d{1,3}\.){3}\d{1,3}/\d{1,2}\b',
    'domain':   r'\b(?:[a-zA-Z0-9](?:[a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}\b',
    'url':      r'https?://[^\s<>"\']{4,300}',
    'email':    r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b',
    'md5':      r'\b[a-fA-F0-9]{32}\b',
    'sha1':     r'\b[a-fA-F0-9]{40}\b',
    'sha256':   r'\b[a-fA-F0-9]{64}\b',
    'cve':      r'CVE-\d{4}-\d{4,7}',
    'btc':      r'\b(?:bc1|[13])[a-zA-HJ-NP-Z0-9]{25,62}\b',
    'eth':      r'\b0x[a-fA-F0-9]{40}\b',
    'mac':      r'\b(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}\b',
    'win_path': r'[a-zA-Z]:\\(?:[^\\/:*?"<>|\r\n]+\\)*[^\\/:*?"<>|\r\n]*',
    'reg_key':  r'HK(?:LM|CU|U|CR|CC)\\[A-Za-z0-9\\_\-. ]{5,200}',
    'phone':    r'\b(?:\+\d{1,3}[-.\s]?)?(?:\d{2,4}[-.\s]?){2,4}\d{3,4}\b',
}
_TLD_WHITELIST = {'png','jpg','jpeg','gif','pdf','exe','dll','txt','log','json','xml','html','htm',
                  'php','asp','aspx','jsp','sh','bat','cmd','ps1','db','dat','bin'}

def _extract_iocs(text: str) -> dict:
    out = {}
    for kind, pat in _IOC_PATTERNS.items():
        matches = set(re.findall(pat, text))
        # 도메인의 경우 파일 확장자 필터링
        if kind == 'domain':
            matches = {m for m in matches if m.split('.')[-1].lower() not in _TLD_WHITELIST}
            matches = {m for m in matches if not re.match(r'^\d+\.\d+\.\d+\.\d+$', m)}
        out[kind] = sorted(matches)[:200]
    return out

@bp.route('/ioc', methods=['GET','POST'])
def ioc_tool():
    result = error = None
    if request.method == 'POST':
        text = (request.form.get('text') or '')
        files = get_files()
        if not text and not files: error = '텍스트 또는 파일 필요'
        else:
            for f in files:
                try:
                    text += '\n' + f.read().decode('utf-8', errors='replace')
                except Exception: pass
            iocs = _extract_iocs(text)
            total = sum(len(v) for v in iocs.values())
            result = {'iocs': iocs, 'total': total, 'text_len': len(text)}
    return render_template('tools/ioc.html', result=result, error=error)


# ============================================================
# /tools/time — 시간 변환기
# ============================================================
@bp.route('/time', methods=['GET','POST'])
def time_tool():
    result = error = None
    if request.method == 'POST':
        val = (request.form.get('value') or '').strip()
        if not val: error = '시각 값 입력'
        else:
            results = []
            # 정수로 파싱
            try: num = int(val)
            except ValueError:
                try: num = int(float(val))
                except ValueError: num = None
            # ISO 문자열로도 시도
            iso_dt = None
            for fmt in ['%Y-%m-%dT%H:%M:%S','%Y-%m-%d %H:%M:%S','%Y-%m-%d',
                        '%Y/%m/%d %H:%M:%S','%d/%m/%Y','%m/%d/%Y']:
                try:
                    iso_dt = _dt.datetime.strptime(val[:19], fmt)
                    break
                except ValueError: pass
            base = iso_dt or _dt.datetime.now(_dt.timezone.utc)
            def fmt_dt(d):
                try: return d.strftime('%Y-%m-%d %H:%M:%S UTC')
                except Exception: return '(범위 초과)'
            if num is not None:
                # 시도: Unix sec / ms / us / ns
                for label, divisor in [('Unix epoch (초)', 1),('Unix epoch (밀리초)', 1e3),
                                       ('Unix epoch (마이크로초)', 1e6),('Unix epoch (나노초)', 1e9)]:
                    try:
                        d = _dt.datetime.fromtimestamp(num/divisor, _dt.timezone.utc)
                        if _dt.datetime(1980,1,1,tzinfo=_dt.timezone.utc) < d < _dt.datetime(2100,1,1,tzinfo=_dt.timezone.utc):
                            results.append({'format': label, 'value': fmt_dt(d), 'raw': num})
                    except Exception: pass
                # Windows FILETIME (100ns since 1601)
                try:
                    d = _dt.datetime(1601,1,1,tzinfo=_dt.timezone.utc) + _dt.timedelta(microseconds=num//10)
                    if d.year < 2100:
                        results.append({'format': 'Windows FILETIME', 'value': fmt_dt(d), 'raw': num})
                except Exception: pass
                # Chrome (us since 1601)
                try:
                    d = _dt.datetime(1601,1,1,tzinfo=_dt.timezone.utc) + _dt.timedelta(microseconds=num)
                    if d.year < 2100:
                        results.append({'format': 'Chrome/Edge epoch (μs since 1601)', 'value': fmt_dt(d), 'raw': num})
                except Exception: pass
                # Cocoa / WebKit (sec since 2001)
                try:
                    d = _dt.datetime(2001,1,1,tzinfo=_dt.timezone.utc) + _dt.timedelta(seconds=num)
                    if d.year < 2100:
                        results.append({'format': 'Cocoa/Safari (sec since 2001)', 'value': fmt_dt(d), 'raw': num})
                except Exception: pass
                # Mozilla (us since 1970)
                try:
                    d = _dt.datetime.fromtimestamp(num/1e6, _dt.timezone.utc)
                    if d.year < 2100:
                        results.append({'format': 'Mozilla Firefox (μs since 1970)', 'value': fmt_dt(d), 'raw': num})
                except Exception: pass
                # DOS Date/Time
                if 0 <= num <= 0xFFFFFFFF:
                    try:
                        dos_date = num & 0xFFFF; dos_time = (num >> 16) & 0xFFFF
                        y = ((dos_date >> 9) & 0x7F) + 1980
                        mo = (dos_date >> 5) & 0xF
                        d = dos_date & 0x1F
                        h = (dos_time >> 11) & 0x1F
                        mi = (dos_time >> 5) & 0x3F
                        s = (dos_time & 0x1F) * 2
                        if 1980 <= y <= 2100 and 1 <= mo <= 12:
                            results.append({'format': 'DOS Date+Time (FAT)',
                                            'value': f'{y}-{mo:02d}-{d:02d} {h:02d}:{mi:02d}:{s:02d}', 'raw': num})
                    except Exception: pass
                # HFS / HFS+ (sec since 1904)
                try:
                    d = _dt.datetime(1904,1,1,tzinfo=_dt.timezone.utc) + _dt.timedelta(seconds=num)
                    if 1980 < d.year < 2100:
                        results.append({'format': 'HFS+ (sec since 1904)', 'value': fmt_dt(d), 'raw': num})
                except Exception: pass
            if iso_dt:
                # 입력이 날짜 → 모든 포맷으로 변환
                d = iso_dt.replace(tzinfo=_dt.timezone.utc)
                unix = int(d.timestamp())
                ft = int((d - _dt.datetime(1601,1,1,tzinfo=_dt.timezone.utc)).total_seconds() * 1e7)
                chrome = int((d - _dt.datetime(1601,1,1,tzinfo=_dt.timezone.utc)).total_seconds() * 1e6)
                cocoa = int((d - _dt.datetime(2001,1,1,tzinfo=_dt.timezone.utc)).total_seconds())
                results.extend([
                    {'format': 'Unix epoch (sec)', 'value': str(unix), 'raw': val},
                    {'format': 'Unix epoch (ms)', 'value': str(unix*1000), 'raw': val},
                    {'format': 'Windows FILETIME', 'value': str(ft), 'raw': val},
                    {'format': 'Chrome epoch', 'value': str(chrome), 'raw': val},
                    {'format': 'Cocoa epoch', 'value': str(cocoa), 'raw': val},
                ])
            result = {'input': val, 'results': results}
    return render_template('tools/time.html', result=result, error=error)


# ============================================================
# /tools/apk — APK 분석기
# ============================================================
def _analyze_apk(data: bytes) -> dict:
    if data[:4] != b'PK\x03\x04': raise ValueError('APK는 ZIP이어야 합니다')
    r = {'size': len(data), 'files': [], 'permissions': [], 'activities': [],
         'services': [], 'receivers': [], 'meta': {}, 'certs': []}
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
        r['file_count'] = len(zf.namelist())
        for zi in zf.infolist()[:500]:
            r['files'].append({'name': zi.filename, 'size': zi.file_size})
        # AndroidManifest.xml 은 binary XML — 헤더만 인식
        if 'AndroidManifest.xml' in zf.namelist():
            manifest = zf.read('AndroidManifest.xml')
            # AXML 시그니처: 0x00080003
            if manifest[:4] == b'\x03\x00\x08\x00':
                r['meta']['manifest_format'] = 'Binary AXML'
                # 단순 문자열 추출
                strings = re.findall(rb'[\x20-\x7E]{4,80}', manifest)
                for s in strings:
                    s = s.decode('latin1', errors='replace')
                    if 'permission.' in s and s not in r['permissions']:
                        r['permissions'].append(s)
                    elif s.endswith('Activity') and s not in r['activities']:
                        r['activities'].append(s)
                    elif s.endswith('Service') and s not in r['services']:
                        r['services'].append(s)
                    elif s.endswith('Receiver') and s not in r['receivers']:
                        r['receivers'].append(s)
        # META-INF/*.RSA / *.DSA — 서명 인증서
        for n in zf.namelist():
            if n.startswith('META-INF/') and (n.endswith('.RSA') or n.endswith('.DSA')):
                cert_data = zf.read(n)
                r['certs'].append({'name': n, 'size': len(cert_data),
                                   'sha256': hashlib.sha256(cert_data).hexdigest()})
        # classes.dex (DEX 헤더)
        if 'classes.dex' in zf.namelist():
            dex = zf.read('classes.dex')
            if dex[:8] == b'dex\n035\x00' or dex[:4] == b'dex\n':
                r['meta']['dex'] = f'DEX 형식 ({len(dex)} bytes)'
                r['meta']['dex_version'] = dex[4:7].decode('latin1', errors='ignore')
        # AndroidManifest.xml content version
        zf.close()
    except Exception as e:
        r['error'] = str(e)
    return r

@bp.route('/apk', methods=['GET','POST'])
def apk_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '.apk 파일 필요'
        else:
            results = []
            for f in files:
                try:
                    r = _analyze_apk(f.read())
                    r['filename'] = f.filename
                    results.append(r)
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/apk.html', result=result, error=error)


# ============================================================
# /tools/hashlookup — 해시 룩업
# ============================================================
# 내장 IoC 해시 데이터베이스 (예시 — 실제로는 NSRL/VT 다운로드 필요)
_KNOWN_HASHES = {
    # SHA-256
    'd41d8cd98f00b204e9800998ecf8427e': ('benign', 'MD5 of empty file'),
    'da39a3ee5e6b4b0d3255bfef95601890afd80709': ('benign', 'SHA-1 of empty file'),
    'e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855': ('benign', 'SHA-256 of empty file'),
}

@bp.route('/hashlookup', methods=['GET','POST'])
def hashlookup_tool():
    result = error = None
    if request.method == 'POST':
        text = (request.form.get('hashes') or '').strip()
        files = get_files()
        hashes = []
        if files:
            for f in files:
                d = f.read()
                hashes.append({
                    'filename': f.filename,
                    'md5': hashlib.md5(d).hexdigest(),
                    'sha1': hashlib.sha1(d).hexdigest(),
                    'sha256': hashlib.sha256(d).hexdigest(),
                    'size': len(d),
                })
        if text:
            for line in text.splitlines():
                line = line.strip().lower()
                if re.match(r'^[a-f0-9]{32}$', line):
                    hashes.append({'filename': '(text)', 'md5': line})
                elif re.match(r'^[a-f0-9]{40}$', line):
                    hashes.append({'filename': '(text)', 'sha1': line})
                elif re.match(r'^[a-f0-9]{64}$', line):
                    hashes.append({'filename': '(text)', 'sha256': line})
        if not hashes: error = '파일 또는 해시 입력'
        else:
            for h in hashes:
                for alg in ['md5','sha1','sha256']:
                    if alg in h and h[alg] in _KNOWN_HASHES:
                        kind, desc = _KNOWN_HASHES[h[alg]]
                        h['known'] = kind
                        h['description'] = desc
                        break
                else:
                    h['known'] = 'unknown'
            result = {'hashes': hashes, 'total': len(hashes),
                      'known_db_size': len(_KNOWN_HASHES)}
    return render_template('tools/hashlookup.html', result=result, error=error)


# ============================================================
# /tools/heif — HEIC/HEIF EXIF
# ============================================================
@bp.route('/heif', methods=['GET','POST'])
def heif_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '.heic / .heif 파일 필요'
        else:
            results = []
            for f in files:
                data = f.read()
                r = {'filename': f.filename, 'size': len(data)}
                # HEIF 시그니처 'ftypheic'/'ftypheix'/'ftypmif1'
                if data[4:8] == b'ftyp':
                    brand = data[8:12].decode('latin1', errors='ignore')
                    r['brand'] = brand
                    r['format'] = {
                        'heic':'HEIC (단일 이미지)','heix':'HEIC 확장',
                        'mif1':'HEIF (단일 이미지)','msf1':'HEIF 시퀀스',
                        'avif':'AVIF (AV1 이미지)',
                    }.get(brand, f'기타 ({brand})')
                # 박스 구조 탐색
                boxes = []
                off = 0
                while off < min(len(data), 1024*1024):
                    if off+8 > len(data): break
                    sz = struct.unpack('>I', data[off:off+4])[0]
                    btype = data[off+4:off+8].decode('latin1', errors='ignore')
                    if sz == 0: break
                    boxes.append({'type': btype, 'size': sz, 'offset': off})
                    if sz == 1:  # extended size
                        sz = struct.unpack('>Q', data[off+8:off+16])[0]
                    off += sz
                    if len(boxes) > 30: break
                r['boxes'] = boxes
                # EXIF 추출 (Pillow에 의존)
                try:
                    from PIL import Image
                    img = Image.open(io.BytesIO(data))
                    r['image_size'] = f'{img.width}x{img.height}'
                    r['image_format'] = img.format
                    if hasattr(img, '_getexif') and img._getexif():
                        r['exif'] = {str(k): str(v)[:200] for k, v in img._getexif().items()}
                except Exception as e:
                    r['pil_note'] = f'Pillow HEIF 미지원 — pillow-heif 설치 필요'
                results.append(r)
            result = {'files': results}
    return render_template('tools/heif.html', result=result, error=error)


# ============================================================
# /tools/memscan — 메모리 덤프 스캔
# ============================================================
def _scan_memory(data: bytes) -> dict:
    r = {'size': len(data), 'iocs': {}, 'processes': [], 'urls': [], 'paths': [],
         'creds': [], 'crypto': []}
    # 청크 단위로 처리
    chunk_size = 1024 * 1024
    full_text = b''
    if len(data) > 200 * 1024 * 1024:
        # 너무 크면 샘플링
        samples = [data[i:i+1024*1024] for i in range(0, len(data), len(data)//20)][:20]
        full_text = b'\n'.join(samples)
    else:
        full_text = data
    # URLs
    r['urls'] = list(set(m.decode('latin1','replace') for m in re.findall(
        rb'https?://[^\x00\s<>"\']{4,200}', full_text)))[:50]
    # 프로세스 이름 (.exe)
    procs = list(set(m.decode('latin1','replace') for m in re.findall(
        rb'[A-Za-z][A-Za-z0-9_\- .]{2,30}\.exe', full_text)))[:80]
    r['processes'] = procs
    # 경로
    paths = list(set(m.decode('latin1','replace') for m in re.findall(
        rb'[A-Z]:\\\\?[A-Za-z0-9_\-. \\\\]{4,150}', full_text)))[:50]
    r['paths'] = paths
    # 자격증명 단서
    cred_patterns = [
        rb'password[\s=:][^\x00\s]{4,40}',
        rb'pwd[\s=:][^\x00\s]{4,40}',
        rb'username[\s=:][^\x00\s]{2,40}',
        rb'token[\s=:][^\x00\s]{10,100}',
    ]
    for pat in cred_patterns:
        for m in re.findall(pat, full_text, re.I)[:20]:
            r['creds'].append(m.decode('latin1', errors='replace'))
    # IoC 추출
    txt = full_text.decode('latin1', errors='replace')
    r['iocs'] = _extract_iocs(txt[:1000000])
    # 메모리 dump 시그니처 인식
    if data[:4] == b'PAGE':
        r['format'] = 'Windows DMP'
    elif data[:4] == b'LiME':
        r['format'] = 'LiME (Linux)'
    elif data[:4] == b'AVML':
        r['format'] = 'AVML (Microsoft)'
    elif data[:4] == b'\x7fELF':
        r['format'] = 'ELF core'
    else:
        r['format'] = 'RAW (분석할 시그니처 없음)'
    return r

@bp.route('/memscan', methods=['GET','POST'])
def memscan_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '메모리 덤프 필요'
        else:
            results = []
            for f in files:
                try:
                    data = f.read()
                    r = _scan_memory(data)
                    r['filename'] = f.filename
                    results.append(r)
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/memscan.html', result=result, error=error)


# ============================================================
# /tools/cuckoo — Cuckoo/CAPE 리포트
# ============================================================
@bp.route('/cuckoo', methods=['GET','POST'])
def cuckoo_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '리포트 JSON 필요'
        else:
            results = []
            for f in files:
                try:
                    report = json.loads(f.read())
                    target = report.get('target', {})
                    info = report.get('info', {})
                    signatures = report.get('signatures', [])
                    network = report.get('network', {})
                    processes = report.get('behavior', {}).get('processes', [])
                    r = {
                        'filename': f.filename,
                        'sample': target.get('file', {}).get('name', ''),
                        'md5': target.get('file', {}).get('md5', ''),
                        'sha256': target.get('file', {}).get('sha256', ''),
                        'score': info.get('score', 0),
                        'category': info.get('category', ''),
                        'started': info.get('started', ''),
                        'ended': info.get('ended', ''),
                        'duration': info.get('duration', 0),
                        'signatures': [
                            {'name': s.get('name', ''),
                             'severity': s.get('severity', 0),
                             'description': s.get('description', '')[:300],
                             'families': s.get('families', [])}
                            for s in signatures
                        ][:50],
                        'process_count': len(processes),
                        'processes': [
                            {'pid': p.get('pid'),
                             'name': p.get('process_name', ''),
                             'parent': p.get('parent_id', 0)}
                            for p in processes
                        ][:30],
                        'hosts': network.get('hosts', [])[:30],
                        'domains': network.get('domains', [])[:30],
                        'http': network.get('http', [])[:30],
                    }
                    results.append(r)
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/cuckoo.html', result=result, error=error)


# ============================================================
# /tools/vol — Volatility 메모리 분석 (스텁)
# ============================================================
@bp.route('/vol', methods=['GET','POST'])
def vol_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '메모리 덤프 필요'
        else:
            try:
                # volatility3 시도
                import volatility3
                error = 'Volatility 3 모듈 호출은 메모리 사용량이 크므로 로컬 스크립트 사용 권장'
                # 실제 구현은 매우 복잡 — 스텁
            except ImportError:
                error = ('Volatility 3 라이브러리 미설치. 다음 명령으로 설치: '
                         'pip install volatility3. 또는 로컬에서 직접 실행하세요.')
    return render_template('tools/vol.html', result=result, error=error)


# ============================================================
# /tools/magic — MIME 시그니처 데이터베이스 (확장)
# ============================================================
# 4000+가 필요하지만 핵심 200+ 시그니처
_MAGIC_DB = [
    # 이미지
    (b'\xFF\xD8\xFF','JPEG','image/jpeg'),(b'\x89PNG\r\n\x1a\n','PNG','image/png'),
    (b'GIF87a','GIF87a','image/gif'),(b'GIF89a','GIF89a','image/gif'),
    (b'BM','BMP','image/bmp'),(b'II*\x00','TIFF LE','image/tiff'),
    (b'MM\x00*','TIFF BE','image/tiff'),(b'RIFF','RIFF (AVI/WAV/WEBP)','application/octet-stream'),
    (b'\x00\x00\x01\x00','ICO','image/x-icon'),(b'\x00\x00\x02\x00','CUR','image/x-icon'),
    (b'\x76\x2F\x31\x01','OpenEXR','image/x-exr'),
    # 비디오
    (b'\x1aE\xdf\xa3','Matroska/WebM','video/x-matroska'),
    (b'\x00\x00\x00\x14ftyp','MP4 (variant A)','video/mp4'),
    (b'\x00\x00\x00\x18ftyp','MP4 (variant B)','video/mp4'),
    (b'\x00\x00\x00\x1cftyp','MP4 (variant C)','video/mp4'),
    (b'\x00\x00\x00\x20ftyp','MP4 (variant D)','video/mp4'),
    (b'FLV','FLV','video/x-flv'),
    # 오디오
    (b'ID3','MP3 (ID3)','audio/mpeg'),(b'\xff\xfb','MP3','audio/mpeg'),
    (b'fLaC','FLAC','audio/flac'),(b'OggS','OGG','audio/ogg'),
    (b'\x4D\x54\x68\x64','MIDI','audio/midi'),
    # 문서
    (b'%PDF','PDF','application/pdf'),
    (b'\xD0\xCF\x11\xE0','OLE2 (DOC/XLS/PPT)','application/vnd.ms-office'),
    (b'PK\x03\x04','ZIP/Office 2007+/Java JAR/APK','application/zip'),
    (b'PK\x05\x06','ZIP (빈)','application/zip'),
    (b'{\\rtf','RTF','application/rtf'),
    # 압축
    (b'Rar!\x1A\x07','RAR','application/vnd.rar'),
    (b'7z\xBC\xAF\x27\x1C','7-Zip','application/x-7z-compressed'),
    (b'\x1F\x8B','GZIP','application/gzip'),
    (b'BZh','BZIP2','application/x-bzip2'),
    (b'\xFD7zXZ','XZ','application/x-xz'),
    (b'ustar','TAR','application/x-tar'),
    # 실행파일
    (b'MZ','PE/EXE/DLL','application/x-msdownload'),
    (b'\x7FELF','ELF','application/x-executable'),
    (b'\xCA\xFE\xBA\xBE','Java class / Mach-O Universal','application/java-vm'),
    (b'\xCE\xFA\xED\xFE','Mach-O 32-bit LE','application/x-mach-binary'),
    (b'\xCF\xFA\xED\xFE','Mach-O 64-bit LE','application/x-mach-binary'),
    (b'\xFE\xED\xFA\xCE','Mach-O 32-bit BE','application/x-mach-binary'),
    (b'\xFE\xED\xFA\xCF','Mach-O 64-bit BE','application/x-mach-binary'),
    (b'#!','Shell 스크립트','text/x-shellscript'),
    (b'\x00asm','WebAssembly','application/wasm'),
    # 데이터베이스
    (b'SQLite format 3\x00','SQLite DB','application/x-sqlite3'),
    (b'\xEF\xCD\xAB\x89','ESE DB (ESEnt)','application/octet-stream'),
    # 포렌식
    (b'regf','Windows Registry','application/octet-stream'),
    (b'ElfFile\x00','EVTX','application/octet-stream'),
    (b'LfLe','EVT (legacy)','application/octet-stream'),
    (b'SCCA','Prefetch (Vista+)','application/octet-stream'),
    (b'MAM\x84','Prefetch 압축 (Win10+)','application/octet-stream'),
    (b'\x4C\x00\x00\x00\x01\x14\x02','LNK','application/x-ms-shortcut'),
    # 디스크 이미지
    (b'EVF\x09\x0D\x0A\xFF\x00','E01 (EnCase)','application/octet-stream'),
    (b'EVF2','Ex01','application/octet-stream'),
    (b'AFF','AFF 이미지','application/octet-stream'),
    (b'AFF4','AFF4','application/octet-stream'),
    (b'vhdxfile','VHDX','application/octet-stream'),
    (b'KDMV','VMDK Sparse','application/octet-stream'),
    (b'QFI\xFB','QCOW2','application/octet-stream'),
    (b'conectix','VHD 푸터','application/octet-stream'),
    # 메모리 덤프
    (b'PAGEDUMP','Windows DMP 32-bit','application/octet-stream'),
    (b'PAGEDU64','Windows DMP 64-bit','application/octet-stream'),
    (b'LiME','LiME 메모리 덤프','application/octet-stream'),
    (b'AVML','AVML 덤프','application/octet-stream'),
    # 암호화
    (b'-----BEGIN','PEM 인증서/키','application/x-pem-file'),
    (b'\x30\x82','DER 인증서/ASN.1','application/x-x509-ca-cert'),
    (b'LUKS\xBA\xBE','LUKS 암호화','application/octet-stream'),
    (b'BMR1','BitLocker','application/octet-stream'),
    # 모바일
    (b'CKMM','iOS .ipa 매니페스트','application/octet-stream'),
    # 폰트
    (b'OTTO','OpenType Font','font/otf'),(b'\x00\x01\x00\x00\x00','TrueType Font','font/ttf'),
    (b'wOFF','WOFF','font/woff'),(b'wOF2','WOFF2','font/woff2'),
    # 텍스트
    (b'\xEF\xBB\xBF','UTF-8 BOM','text/plain'),
    (b'\xFE\xFF','UTF-16 BE BOM','text/plain'),
    (b'\xFF\xFE','UTF-16 LE BOM','text/plain'),
    # 기타
    (b'CWS','Adobe Flash SWF (압축)','application/x-shockwave-flash'),
    (b'FWS','Adobe Flash SWF','application/x-shockwave-flash'),
    (b'8BPS','Photoshop PSD','image/vnd.adobe.photoshop'),
    (b'TPL','TPL (게임 텍스처)','application/octet-stream'),
    (b'WAD!','DOOM WAD','application/octet-stream'),
    (b'\xfb\xff\xff\xff','Bitcoin Block','application/octet-stream'),
]

@bp.route('/magic', methods=['GET','POST'])
def magic_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '파일 필요'
        else:
            results = []
            for f in files:
                head = f.read(256)
                matches = []
                for sig, label, mime in _MAGIC_DB:
                    if head.startswith(sig):
                        matches.append({
                            'sig': sig.hex().upper(),
                            'label': label.decode('latin1') if isinstance(label, bytes) else label,
                            'mime': mime,
                        })
                results.append({
                    'filename': f.filename, 'size_head': len(head),
                    'matches': matches,
                    'hex_preview': head[:64].hex(),
                })
            result = {'files': results, 'db_size': len(_MAGIC_DB)}
    return render_template('tools/magic.html', result=result, error=error)


# ============================================================
# /tools/docker — Docker 이미지 레이어
# ============================================================
@bp.route('/docker', methods=['GET','POST'])
def docker_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '.tar 파일 필요'
        else:
            results = []
            for f in files:
                try:
                    import tarfile
                    tf = tarfile.open(fileobj=io.BytesIO(f.read()))
                    members = tf.getnames()
                    layers = [m for m in members if m.endswith('/layer.tar') or m.endswith('.tar.gz')]
                    manifest = None
                    config = None
                    for m in members:
                        if m == 'manifest.json':
                            manifest = json.loads(tf.extractfile(m).read())
                        elif m.endswith('.json') and 'config' in m.lower() and len(m) > 12:
                            try: config = json.loads(tf.extractfile(m).read())
                            except Exception: pass
                    r = {
                        'filename': f.filename,
                        'member_count': len(members),
                        'layers': layers,
                        'manifest': manifest,
                        'config_keys': list(config.keys()) if config else [],
                        'config_history': (config or {}).get('history', [])[:30],
                        'image_id': (manifest[0] if manifest else {}).get('Config', ''),
                        'tags': (manifest[0] if manifest else {}).get('RepoTags', []),
                    }
                    if config:
                        r['os'] = config.get('os', '')
                        r['architecture'] = config.get('architecture', '')
                        cfg = config.get('config', {})
                        r['env'] = cfg.get('Env', [])
                        r['cmd'] = cfg.get('Cmd', [])
                        r['entrypoint'] = cfg.get('Entrypoint', [])
                        r['exposed_ports'] = list((cfg.get('ExposedPorts') or {}).keys())
                        r['volumes'] = list((cfg.get('Volumes') or {}).keys())
                    tf.close()
                    results.append(r)
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/docker.html', result=result, error=error)


# ============================================================
# /tools/hex — Hex Viewer + 검색
# ============================================================
@bp.route('/hex', methods=['GET','POST'])
def hex_tool():
    result = error = None
    if request.method == 'POST':
        f = request.files.get('file')
        offset = int(request.form.get('offset', 0) or 0)
        length = min(int(request.form.get('length', 1024) or 1024), 65536)
        search = (request.form.get('search') or '').strip()
        if not f or not f.filename: error = '파일 필요'
        else:
            data = f.read()
            chunk = data[offset:offset+length]
            # 헥스 라인 16바이트씩
            lines = []
            for i in range(0, len(chunk), 16):
                b = chunk[i:i+16]
                hex_part = ' '.join(f'{x:02X}' for x in b)
                ascii_part = ''.join(chr(x) if 32 <= x < 127 else '.' for x in b)
                lines.append({
                    'offset': f'{offset+i:08X}',
                    'hex': hex_part.ljust(48),
                    'ascii': ascii_part,
                })
            search_results = []
            if search:
                try:
                    if search.startswith('0x'):
                        pat = bytes.fromhex(search[2:])
                    else:
                        pat = search.encode('utf-8')
                    pos = 0
                    while pos < len(data) and len(search_results) < 100:
                        idx = data.find(pat, pos)
                        if idx < 0: break
                        search_results.append({
                            'offset': f'{idx:08X}',
                            'context': data[max(0,idx-8):idx+len(pat)+8].hex().upper(),
                        })
                        pos = idx + 1
                except Exception as e:
                    error = f'검색 오류: {e}'
            result = {
                'filename': f.filename, 'file_size': len(data),
                'offset': offset, 'length': len(chunk),
                'lines': lines, 'search': search, 'search_results': search_results,
            }
    return render_template('tools/hex.html', result=result, error=error)


# ============================================================
# /tools/cidr — CIDR 계산기
# ============================================================
@bp.route('/cidr', methods=['GET','POST'])
def cidr_tool():
    result = error = None
    if request.method == 'POST':
        cidr = (request.form.get('cidr') or '').strip()
        if not cidr: error = 'CIDR 표기 입력 (예: 192.168.1.0/24)'
        else:
            try:
                import ipaddress
                net = ipaddress.ip_network(cidr, strict=False)
                result = {
                    'input': cidr,
                    'network': str(net.network_address),
                    'netmask': str(net.netmask),
                    'broadcast': str(net.broadcast_address) if hasattr(net, 'broadcast_address') else '',
                    'wildcard': str(net.hostmask) if hasattr(net, 'hostmask') else '',
                    'first': str(net.network_address + 1) if net.num_addresses > 2 else str(net.network_address),
                    'last': str(net.broadcast_address - 1) if hasattr(net, 'broadcast_address') and net.num_addresses > 2 else '',
                    'total': net.num_addresses,
                    'usable': max(0, net.num_addresses - 2) if net.version == 4 else net.num_addresses,
                    'prefix': net.prefixlen,
                    'is_private': net.is_private,
                    'is_global': net.is_global,
                    'is_loopback': net.is_loopback,
                    'is_link_local': net.is_link_local,
                    'is_multicast': net.is_multicast,
                    'version': net.version,
                }
                # 첫 8개 호스트
                hosts = []
                for i, ip in enumerate(net.hosts() if net.num_addresses > 2 else [net.network_address]):
                    if i >= 8: break
                    hosts.append(str(ip))
                result['sample_hosts'] = hosts
            except Exception as e: error = str(e)
    return render_template('tools/cidr.html', result=result, error=error)


# ============================================================
# /tools/convert — JSON/XML/YAML 변환
# ============================================================
@bp.route('/convert', methods=['GET','POST'])
def convert_tool():
    result = error = None
    text = ''; mode = ''
    if request.method == 'POST':
        text = (request.form.get('text') or '').strip()
        mode = request.form.get('mode', 'pretty_json')
        if not text: error = '텍스트 입력 필요'
        else:
            try:
                obj = None
                input_format = '?'
                # 자동 감지
                if text.startswith('{') or text.startswith('['):
                    obj = json.loads(text); input_format = 'JSON'
                elif text.startswith('<'):
                    try:
                        import xml.etree.ElementTree as ET
                        root = ET.fromstring(text)
                        def xml_to_dict(elem):
                            d = {}
                            d.update(('@'+k, v) for k, v in elem.attrib.items())
                            for child in elem:
                                cd = xml_to_dict(child)
                                if child.tag in d:
                                    if not isinstance(d[child.tag], list):
                                        d[child.tag] = [d[child.tag]]
                                    d[child.tag].append(cd)
                                else:
                                    d[child.tag] = cd
                            if elem.text and elem.text.strip(): d['#text'] = elem.text.strip()
                            return d
                        obj = {root.tag: xml_to_dict(root)}
                        input_format = 'XML'
                    except Exception: pass
                else:
                    try:
                        import yaml as _yaml
                        obj = _yaml.safe_load(text); input_format = 'YAML'
                    except Exception: pass
                if obj is None:
                    error = '입력을 JSON/XML/YAML로 파싱할 수 없습니다'
                else:
                    output = ''
                    if mode == 'pretty_json':
                        output = json.dumps(obj, indent=2, ensure_ascii=False, default=str)
                    elif mode == 'compact_json':
                        output = json.dumps(obj, ensure_ascii=False, default=str)
                    elif mode == 'yaml':
                        try:
                            import yaml as _yaml
                            output = _yaml.dump(obj, allow_unicode=True, default_flow_style=False)
                        except ImportError:
                            output = json.dumps(obj, indent=2, ensure_ascii=False, default=str)
                            error = 'pyyaml 미설치, JSON으로 대체'
                    result = {'input_format': input_format, 'output': output, 'mode': mode}
            except Exception as e: error = str(e)
    return render_template('tools/convert.html', result=result, error=error,
                           text=text, mode=mode)


# ============================================================
# /tools/regex — 정규식 테스터
# ============================================================
@bp.route('/regex', methods=['GET','POST'])
def regex_tool():
    result = error = None
    pattern = text = ''
    flags = []
    if request.method == 'POST':
        pattern = request.form.get('pattern', '')
        text = request.form.get('text', '')
        flags = request.form.getlist('flags')
        if not pattern: error = '정규식 입력'
        elif not text: error = '검사 텍스트 입력'
        else:
            try:
                fl = 0
                if 'i' in flags: fl |= re.IGNORECASE
                if 'm' in flags: fl |= re.MULTILINE
                if 's' in flags: fl |= re.DOTALL
                if 'x' in flags: fl |= re.VERBOSE
                regex = re.compile(pattern, fl)
                matches = []
                for m in regex.finditer(text):
                    matches.append({
                        'match': m.group(0),
                        'groups': [m.group(i) for i in range(1, (m.lastindex or 0)+1)],
                        'span': [m.start(), m.end()],
                        'named': dict(m.groupdict()),
                    })
                    if len(matches) >= 1000: break
                # substitution preview
                sub_repl = request.form.get('replace', '')
                substituted = regex.sub(sub_repl, text) if sub_repl else None
                result = {
                    'pattern': pattern, 'flags': flags,
                    'match_count': len(matches),
                    'matches': matches[:500],
                    'substituted': substituted,
                }
            except Exception as e: error = f'정규식 오류: {e}'
    return render_template('tools/regex.html', result=result, error=error,
                           pattern=pattern, text=text, flags=flags)


# ============================================================
# /tools/jsdeobf — JS 디오브푸스케이션
# ============================================================
@bp.route('/jsdeobf', methods=['GET','POST'])
def jsdeobf_tool():
    result = error = None
    text = ''
    if request.method == 'POST':
        text = (request.form.get('text') or '').strip()
        f = request.files.get('file')
        if not text and f and f.filename:
            text = f.read().decode('utf-8', errors='replace')
        if not text: error = 'JS 코드 필요'
        else:
            steps = []
            current = text
            # eval() 펴기
            eval_pat = re.compile(r'eval\s*\(\s*["\']([^"\']+)["\']\s*\)')
            for m in eval_pat.finditer(current):
                current = current.replace(m.group(0), m.group(1))
                steps.append({'step': 'eval() 펴기', 'before': m.group(0)[:80], 'after': m.group(1)[:120]})
            # \xNN 헥스 이스케이프
            hex_pat = re.compile(r'\\x([0-9a-fA-F]{2})')
            def hex_repl(m):
                try: return chr(int(m.group(1), 16))
                except Exception: return m.group(0)
            new = hex_pat.sub(hex_repl, current)
            if new != current:
                steps.append({'step': '\\xNN 디코드', 'before': '', 'after': ''})
                current = new
            # \uNNNN 유니코드 이스케이프
            uni_pat = re.compile(r'\\u([0-9a-fA-F]{4})')
            def uni_repl(m):
                try: return chr(int(m.group(1), 16))
                except Exception: return m.group(0)
            new = uni_pat.sub(uni_repl, current)
            if new != current:
                steps.append({'step': '\\uNNNN 디코드', 'before': '', 'after': ''})
                current = new
            # btoa / atob (base64)
            atob_pat = re.compile(r'atob\s*\(\s*["\']([A-Za-z0-9+/=]+)["\']\s*\)')
            for m in atob_pat.finditer(current):
                try:
                    decoded = base64.b64decode(m.group(1)).decode('utf-8', errors='replace')
                    current = current.replace(m.group(0), f'"{decoded}"')
                    steps.append({'step': 'atob() Base64', 'before': m.group(1)[:60], 'after': decoded[:120]})
                except Exception: pass
            # String.fromCharCode
            fcc_pat = re.compile(r'String\.fromCharCode\s*\(([\d,\s]+)\)')
            for m in fcc_pat.finditer(current):
                try:
                    nums = [int(x.strip()) for x in m.group(1).split(',')]
                    chars = ''.join(chr(n) for n in nums if 0 < n < 65536)
                    current = current.replace(m.group(0), f'"{chars}"')
                    steps.append({'step': 'fromCharCode', 'before': m.group(1)[:60], 'after': chars[:120]})
                except Exception: pass
            # 문자열 연결
            cat_pat = re.compile(r'"([^"]{1,80})"\s*\+\s*"([^"]{1,80})"')
            for _ in range(30):
                new = cat_pat.sub(lambda m: f'"{m.group(1)}{m.group(2)}"', current)
                if new == current: break
                current = new
            steps.append({'step': '문자열 연결 정리', 'before': '', 'after': ''})
            # IoC 추출
            iocs = _extract_iocs(current)
            result = {'original': text[:5000], 'result': current[:50000],
                      'steps': steps, 'iocs': iocs}
    return render_template('tools/jsdeobf.html', result=result, error=error, text=text)


# ============================================================
# /tools/wordlist — Wordlist 생성
# ============================================================
def _gen_wordlist(words: list, options: dict) -> list:
    out = set(words)
    # Leetspeak
    if options.get('leet'):
        leet_map = {'a':'@','e':'3','i':'1','o':'0','s':'$','t':'7','g':'9','b':'8'}
        for w in list(out)[:100]:
            new = ''.join(leet_map.get(c.lower(), c) for c in w)
            out.add(new)
    # 대소문자
    if options.get('case'):
        for w in list(out)[:500]:
            out.add(w.lower()); out.add(w.upper()); out.add(w.capitalize())
    # 연도 접미
    if options.get('years'):
        years = range(int(options.get('year_from', 2020)), int(options.get('year_to', 2026)) + 1)
        for w in list(out)[:500]:
            for y in years:
                out.add(f'{w}{y}'); out.add(f'{w}{y%100:02d}')
    # 숫자 접미
    if options.get('numbers'):
        for w in list(out)[:300]:
            for n in range(int(options.get('num_max', 100))):
                out.add(f'{w}{n}')
    # 특수문자
    if options.get('symbols'):
        for w in list(out)[:300]:
            for s in ['!','@','#','$','*','?']:
                out.add(f'{w}{s}')
    # 길이 필터
    min_len = int(options.get('min_len', 1))
    max_len = int(options.get('max_len', 30))
    return sorted([w for w in out if min_len <= len(w) <= max_len])[:10000]

@bp.route('/wordlist', methods=['GET','POST'])
def wordlist_tool():
    result = error = None
    if request.method == 'POST':
        seed_text = (request.form.get('seeds') or '').strip()
        if not seed_text: error = '시드 단어 입력'
        else:
            seeds = [w.strip() for w in re.split(r'[\s,\n]+', seed_text) if w.strip()]
            opts = {
                'leet': 'leet' in request.form,
                'case': 'case' in request.form,
                'years': 'years' in request.form,
                'numbers': 'numbers' in request.form,
                'symbols': 'symbols' in request.form,
                'year_from': request.form.get('year_from', 2020),
                'year_to': request.form.get('year_to', 2026),
                'num_max': request.form.get('num_max', 100),
                'min_len': request.form.get('min_len', 4),
                'max_len': request.form.get('max_len', 20),
            }
            wl = _gen_wordlist(seeds, opts)
            result = {'words': wl, 'count': len(wl), 'options': opts}
    return render_template('tools/wordlist.html', result=result, error=error)


# ============================================================
# /tools/spreadsheet — CSV/Excel 뷰어
# ============================================================
@bp.route('/spreadsheet', methods=['GET','POST'])
def spreadsheet_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = '파일 필요'
        else:
            results = []
            for f in files:
                data = f.read()
                try:
                    if f.filename.lower().endswith(('.csv','.tsv','.txt')):
                        sep = '\t' if f.filename.endswith('.tsv') else None
                        import csv as _csv
                        text = data.decode('utf-8', errors='replace')
                        if sep is None:
                            sniffer = _csv.Sniffer()
                            try: sep = sniffer.sniff(text[:2048]).delimiter
                            except Exception: sep = ','
                        reader = _csv.reader(io.StringIO(text), delimiter=sep)
                        rows = list(reader)
                        headers = rows[0] if rows else []
                        data_rows = rows[1:1000]
                        results.append({
                            'filename': f.filename, 'format': 'CSV',
                            'delimiter': sep,
                            'rows_total': len(rows) - 1,
                            'cols': len(headers),
                            'headers': headers,
                            'rows': data_rows,
                        })
                    elif f.filename.lower().endswith(('.xlsx','.xls')):
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
                            sheet_data = []
                            for sheet in wb.sheetnames[:10]:
                                ws = wb[sheet]
                                rows = list(ws.iter_rows(values_only=True))
                                sheet_data.append({
                                    'name': sheet, 'rows': len(rows),
                                    'cols': max((len(r) for r in rows), default=0),
                                    'headers': list(rows[0]) if rows else [],
                                    'data': [list(r) for r in rows[1:1000]],
                                })
                            results.append({
                                'filename': f.filename, 'format': 'XLSX',
                                'sheets': sheet_data, 'sheet_count': len(wb.sheetnames),
                            })
                        except ImportError:
                            results.append({'filename': f.filename, 'error': 'openpyxl 라이브러리 미설치'})
                except Exception as e:
                    results.append({'filename': f.filename, 'error': str(e)})
            result = {'files': results}
    return render_template('tools/spreadsheet.html', result=result, error=error)


# ============================================================
# /tools/textdiff — 텍스트 diff
# ============================================================
@bp.route('/textdiff', methods=['GET','POST'])
def textdiff_tool():
    result = error = None
    a = b = ''
    if request.method == 'POST':
        a = request.form.get('text_a', '')
        b = request.form.get('text_b', '')
        if not a or not b: error = '두 텍스트 모두 입력'
        else:
            import difflib
            differ = difflib.unified_diff(
                a.splitlines(keepends=False),
                b.splitlines(keepends=False),
                fromfile='A', tofile='B', lineterm='', n=3)
            diff_lines = list(differ)
            ratio = difflib.SequenceMatcher(None, a, b).ratio()
            result = {
                'diff': diff_lines, 'similarity': round(ratio * 100, 2),
                'lines_a': a.count('\n')+1, 'lines_b': b.count('\n')+1,
                'len_a': len(a), 'len_b': len(b),
            }
    return render_template('tools/textdiff.html', result=result, error=error, a=a, b=b)


# ============================================================
# /tools/cve — CVE 검색
# ============================================================
# 내장 CVE DB (실제로는 NVD JSON 피드를 캐싱)
_CVE_DB = {
    'CVE-2021-44228': {'name':'Log4Shell','severity':'10.0 CRITICAL',
        'description':'Apache Log4j 2 JNDI 원격 코드 실행',
        'products':['Apache Log4j 2'],'date':'2021-12-10'},
    'CVE-2017-0144': {'name':'EternalBlue','severity':'8.1 HIGH',
        'description':'Windows SMBv1 원격 코드 실행',
        'products':['Microsoft Windows'],'date':'2017-03-14'},
    'CVE-2014-0160': {'name':'Heartbleed','severity':'7.5 HIGH',
        'description':'OpenSSL Heartbeat 정보 누출',
        'products':['OpenSSL 1.0.1'],'date':'2014-04-07'},
    'CVE-2014-6271': {'name':'Shellshock','severity':'9.8 CRITICAL',
        'description':'Bash 환경변수 코드 실행',
        'products':['GNU Bash'],'date':'2014-09-24'},
    'CVE-2017-5638': {'name':'Apache Struts RCE','severity':'10.0 CRITICAL',
        'description':'Struts2 Content-Type RCE','products':['Apache Struts'],'date':'2017-03-06'},
    'CVE-2019-0708': {'name':'BlueKeep','severity':'9.8 CRITICAL',
        'description':'Windows RDP 원격 코드 실행','products':['Windows 7/2008'],'date':'2019-05-14'},
    'CVE-2020-1472': {'name':'Zerologon','severity':'10.0 CRITICAL',
        'description':'Netlogon 권한 상승','products':['Windows Server'],'date':'2020-08-11'},
    'CVE-2021-34527': {'name':'PrintNightmare','severity':'8.8 HIGH',
        'description':'Windows Print Spooler RCE','products':['Windows'],'date':'2021-07-01'},
    'CVE-2022-30190': {'name':'Follina','severity':'7.8 HIGH',
        'description':'Microsoft Office MSDT URL 핸들러 RCE',
        'products':['Microsoft Office'],'date':'2022-05-30'},
    'CVE-2023-23397': {'name':'Outlook NTLM Leak','severity':'9.8 CRITICAL',
        'description':'Outlook 권한 상승 — NTLM 해시 누출',
        'products':['Microsoft Outlook'],'date':'2023-03-14'},
}

@bp.route('/cve', methods=['GET','POST'])
def cve_tool():
    result = error = None
    if request.method == 'POST':
        q = (request.form.get('query') or '').strip().upper()
        if not q: error = 'CVE ID 또는 키워드 입력'
        else:
            matches = []
            # 정확한 CVE
            if q in _CVE_DB:
                matches.append({'id': q, **_CVE_DB[q]})
            else:
                # 키워드 검색
                for cve_id, info in _CVE_DB.items():
                    if (q in cve_id or q.lower() in info['name'].lower()
                            or q.lower() in info['description'].lower()
                            or any(q.lower() in p.lower() for p in info['products'])):
                        matches.append({'id': cve_id, **info})
            result = {'query': q, 'matches': matches, 'db_size': len(_CVE_DB)}
    return render_template('tools/cve.html', result=result, error=error)


# ============================================================
# /tools/phash — 이미지 Perceptual Hash
# ============================================================
def _phash(img_data: bytes, size=8) -> str:
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(img_data)).convert('L').resize((size*4, size*4), Image.Resampling.LANCZOS)
        # 평균 해시 (aHash)
        pixels = list(img.getdata())
        avg = sum(pixels) / len(pixels)
        bits = ''.join('1' if p > avg else '0' for p in pixels)
        # 16진수로 변환
        h = hex(int(bits, 2))[2:].zfill(size*size//4)
        return h
    except Exception as e: return f'error:{e}'

def _hamming(a: str, b: str) -> int:
    if len(a) != len(b): return -1
    try:
        return bin(int(a, 16) ^ int(b, 16)).count('1')
    except Exception: return -1

@bp.route('/phash', methods=['GET','POST'])
def phash_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if len(files) < 1: error = '이미지 필요'
        else:
            hashes = []
            for f in files:
                data = f.read()
                h = _phash(data)
                hashes.append({'filename': f.filename, 'size': len(data),
                               'phash': h,
                               'sha256': hashlib.sha256(data).hexdigest()})
            # 두 개 이상이면 페어별 거리
            pairs = []
            if len(hashes) >= 2:
                for i in range(len(hashes)):
                    for j in range(i+1, len(hashes)):
                        d = _hamming(hashes[i]['phash'], hashes[j]['phash'])
                        if d >= 0:
                            pairs.append({
                                'a': hashes[i]['filename'], 'b': hashes[j]['filename'],
                                'distance': d,
                                'similarity': round((1 - d/64) * 100, 1) if d >= 0 else 0,
                            })
                pairs.sort(key=lambda x: x['distance'])
            result = {'hashes': hashes, 'pairs': pairs[:50]}
    return render_template('tools/phash.html', result=result, error=error)


# ============================================================
# /tools/dmesg — Linux dmesg / journalctl 분석
# ============================================================
@bp.route('/dmesg', methods=['GET','POST'])
def dmesg_tool():
    result = error = None
    if request.method == 'POST':
        f = request.files.get('file')
        text = (request.form.get('text') or '').strip()
        if f and f.filename:
            text = f.read().decode('utf-8', errors='replace')
        if not text: error = '로그 텍스트 또는 파일 필요'
        else:
            lines = text.splitlines()[:10000]
            events = []
            keywords = Counter()
            errors = warns = 0
            for i, line in enumerate(lines):
                ev = {'line': i+1, 'raw': line[:300]}
                # [12345.678] timestamp 또는 ISO datetime
                m = re.match(r'\[([\d.]+)\]\s+(.*)', line)
                if m:
                    ev['boot_time'] = m.group(1)
                    ev['msg'] = m.group(2)
                # journalctl: Jun 01 12:34:56 hostname process[pid]: msg
                m = re.match(r'(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(\S+?)(?:\[(\d+)\])?:\s+(.*)', line)
                if m:
                    ev['timestamp'] = m.group(1)
                    ev['host'] = m.group(2)
                    ev['process'] = m.group(3)
                    ev['pid'] = m.group(4)
                    ev['msg'] = m.group(5)
                # 심각도
                lo = line.lower()
                if any(w in lo for w in ['error','failed','denied','panic','oops','segfault']):
                    ev['severity'] = 'error'; errors += 1
                elif any(w in lo for w in ['warning','warn','deprecated']):
                    ev['severity'] = 'warn'; warns += 1
                else:
                    ev['severity'] = 'info'
                # 카테고리 키워드
                for cat in ['usb','disk','network','memory','cpu','firewall','iptables',
                            'sshd','sudo','kernel','systemd','docker','selinux','apparmor']:
                    if cat in lo:
                        keywords[cat] += 1
                        ev['category'] = cat
                        break
                events.append(ev)
            result = {
                'total': len(events),
                'events': events[:500],
                'errors': errors, 'warnings': warns,
                'top_categories': keywords.most_common(),
            }
    return render_template('tools/dmesg.html', result=result, error=error)


# ============================================================
# /tools/ios-backup — iOS Manifest.db
# ============================================================
@bp.route('/ios-backup', methods=['GET','POST'])
def ios_backup_tool():
    result = error = None
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename: error = 'Manifest.db 파일 필요'
        else:
            data = f.read()
            if data[:16] != b'SQLite format 3\x00':
                error = 'SQLite 파일이 아님'
            else:
                import sqlite3 as _sqlite3, tempfile
                tf = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
                tf.write(data); tf.close()
                try:
                    con = _sqlite3.connect(f'file:{tf.name}?mode=ro', uri=True)
                    cur = con.cursor()
                    # Files 테이블 확인
                    files = []
                    domains = Counter()
                    apps = Counter()
                    try:
                        for row in cur.execute(
                                'SELECT fileID, domain, relativePath, flags FROM Files LIMIT 5000'):
                            files.append({
                                'fileID': row[0], 'domain': row[1] or '',
                                'path': row[2] or '', 'flags': row[3],
                            })
                            d = row[1] or ''
                            domains[d] += 1
                            if d.startswith('AppDomain-'):
                                apps[d.replace('AppDomain-','')] += 1
                    except Exception as e:
                        error = f'Files 테이블 조회 실패: {e}'
                    result = {
                        'filename': f.filename, 'size': len(data),
                        'total_files': len(files),
                        'top_domains': domains.most_common(30),
                        'top_apps': apps.most_common(50),
                        'sample_files': files[:200],
                    }
                    con.close()
                finally: os.unlink(tf.name)
    return render_template('tools/ios_backup.html', result=result, error=error)


# ============================================================
# /tools/whatsapp — WhatsApp DB
# ============================================================
@bp.route('/whatsapp', methods=['GET','POST'])
def whatsapp_tool():
    result = error = None
    if request.method == 'POST':
        f = request.files.get('file')
        key = (request.form.get('key') or '').strip()
        if not f or not f.filename: error = 'msgstore.db 또는 .crypt 파일 필요'
        else:
            data = f.read()
            r = {'filename': f.filename, 'size': len(data)}
            if data[:16] == b'SQLite format 3\x00':
                # 평문 SQLite
                import sqlite3 as _sqlite3, tempfile
                tf = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
                tf.write(data); tf.close()
                try:
                    con = _sqlite3.connect(f'file:{tf.name}?mode=ro', uri=True)
                    cur = con.cursor()
                    r['format'] = 'SQLite (평문)'
                    tables = [t[0] for t in cur.execute(
                        "SELECT name FROM sqlite_master WHERE type='table'")]
                    r['tables'] = tables
                    if 'messages' in tables:
                        msgs = list(cur.execute(
                            "SELECT _id, key_remote_jid, key_from_me, timestamp, data, "
                            "media_wa_type FROM messages ORDER BY timestamp DESC LIMIT 100"))
                        r['messages'] = [
                            {'id': m[0], 'chat': m[1], 'from_me': bool(m[2]),
                             'time': _dt.datetime.fromtimestamp(m[3]/1000).isoformat() if m[3] else '',
                             'text': (m[4] or '')[:300], 'type': m[5]} for m in msgs]
                        r['msg_count'] = list(cur.execute("SELECT COUNT(*) FROM messages"))[0][0]
                    con.close()
                finally: os.unlink(tf.name)
            elif data[:4] == b'\x00\x01\x02\x03' or data[:1] == b'\x02':
                r['format'] = 'WhatsApp Crypt14/15 (암호화)'
                r['note'] = ('암호화된 백업입니다. 디크립트를 위해 32바이트 키가 필요합니다.\n'
                            '키 위치: /data/data/com.whatsapp/files/key 또는 \n'
                            'sdcard/Android/data/com.whatsapp/files/decrypted_backup.key')
                if key and len(key) == 64:
                    r['note'] += '\n(키 디크립트는 별도 도구 사용 권장: github.com/MaxiHuHe04/Whatsapp-Crypt14-Decrypter)'
            else:
                r['format'] = '알 수 없는 형식'
                r['head_hex'] = data[:16].hex()
            result = r
    return render_template('tools/whatsapp.html', result=result, error=error)


# ============================================================
# /tools/telegram — Telegram tdata
# ============================================================
@bp.route('/telegram', methods=['GET','POST'])
def telegram_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = 'Telegram tdata 파일 필요'
        else:
            results = []
            for f in files:
                data = f.read()
                r = {'filename': f.filename, 'size': len(data)}
                # TDF$ 시그니처
                if data[:4] == b'TDF$':
                    r['format'] = 'Telegram TDF$ (암호화 가능)'
                    if len(data) >= 12:
                        version = struct.unpack('<I', data[4:8])[0]
                        length = struct.unpack('<I', data[8:12])[0]
                        r['version'] = version
                        r['data_length'] = length
                        r['md5_check'] = hashlib.md5(data[:-16]).hexdigest() == data[-16:].hex()
                else:
                    r['format'] = '알 수 없는 형식'
                    r['head_hex'] = data[:16].hex()
                # 키 데이터 단서 (UTF-16 문자열)
                strings = []
                for m in re.finditer(rb'(?:[\x20-\x7E]\x00){5,40}', data):
                    s = m.group().decode('utf-16-le', errors='replace').rstrip('\x00')
                    if s and s.isprintable():
                        strings.append(s)
                r['utf16_strings'] = list(set(strings))[:30]
                results.append(r)
            result = {'files': results}
    return render_template('tools/telegram.html', result=result, error=error)


# ============================================================
# /tools/pst — Outlook PST/OST 헤더
# ============================================================
@bp.route('/pst', methods=['GET','POST'])
def pst_tool():
    result = error = None
    if request.method == 'POST':
        files = get_files()
        if not files: error = 'PST/OST 파일 필요'
        else:
            results = []
            for f in files:
                head = f.read(512)
                r = {'filename': f.filename}
                if head[:4] != b'!BDN':
                    r['error'] = '!BDN 시그니처 없음 — PST/OST 아님'
                else:
                    ver = struct.unpack('<H', head[8:10])[0]
                    client_ver = struct.unpack('<H', head[10:12])[0]
                    platform_create = head[12]
                    platform_access = head[13]
                    crypt = head[0x180]
                    FORMATS = {14:'ANSI (PST 97-2002)',15:'ANSI',23:'Unicode (PST 2003+)',
                               36:'Unicode 4K'}
                    CRYPT = {0:'None',1:'Permute (NDB Crypt)',2:'Cyclic (Office 2003+)'}
                    r['signature'] = '!BDN'
                    r['format'] = FORMATS.get(ver, f'Unknown ({ver})')
                    r['version_client'] = client_ver
                    r['encryption'] = CRYPT.get(crypt, f'Unknown ({crypt})')
                    r['platform_create'] = platform_create
                    r['platform_access'] = platform_access
                    r['note'] = '전체 분석은 libpff 또는 readpst 도구 사용을 권장합니다.'
                results.append(r)
            result = {'files': results}
    return render_template('tools/pst.html', result=result, error=error)
